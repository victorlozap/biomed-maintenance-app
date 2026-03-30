import openpyxl
import os

folder = r'd:\VICTOR LOPEZ\CLAUDE CODE\prueba\formatos'
output_file = r'd:\VICTOR LOPEZ\CLAUDE CODE\prueba\format_analysis.txt'

files = sorted([f for f in os.listdir(folder) if f.endswith('.xlsx')])

with open(output_file, 'w', encoding='utf-8') as out:
    for fname in files:
        fpath = os.path.join(folder, fname)
        out.write('='*65 + '\n')
        out.write('FILE: ' + fname + '\n')
        out.write('='*65 + '\n')
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for shname in wb.sheetnames[:2]:
                out.write(f'  [SHEET: {shname}]\n')
                ws = wb[shname]
                for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
                    vals = [str(v).strip() for v in row if v is not None and str(v).strip() not in ('None', '')]
                    if vals:
                        out.write('    ' + ' | '.join(vals[:10]) + '\n')
        except Exception as e:
            out.write(f'  ERROR: {e}\n')
        out.write('\n')

print('Done! Output written to format_analysis.txt')
