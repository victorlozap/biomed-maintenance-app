import openpyxl
import os

folder = r'd:\VICTOR LOPEZ\CLAUDE CODE\prueba\formatos'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]

for fname in files[:5]:
    fpath = os.path.join(folder, fname)
    print('='*60)
    print('FILE:', fname[:55])
    print('='*60)
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for shname in wb.sheetnames[:2]:
            print(f'  SHEET: {shname}')
            ws = wb[shname]
            for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
                vals = [str(v).strip() for v in row if v is not None and str(v).strip() not in ('None', '')]
                if vals:
                    print('    ' + ' | '.join(vals[:8]))
    except Exception as e:
        print(f'  ERROR: {e}')
    print()
